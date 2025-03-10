import mvcm
import os
import tempfile

def printResponseError(response):
    # Print all details about the response
    print(f"Status Code: {response.status_code}")
    print(f"Headers: {response.headers}")
    print(f"Content: {response.content}")
    print(f"Text: {response.text}")
    print(f"JSON: {response.json() if response.headers.get('Content-Type') == 'application/json' else 'Not a JSON response'}")
    print(f"URL: {response.url}")
    print(f"Elapsed Time: {response.elapsed}")
    print(f"Request Headers: {response.request.headers}")
    print(f"Request Method: {response.request.method}")
    print(f"Request URL: {response.request.url}")

class BusinessController:
    def __init__(self, mvcm_instance):
        self.mvcm = mvcm_instance

    def connect(self, hostname, username, password):
        try:
            self.mvcm.connect(hostname, username, password)
        except Exception:
            pass
        
    def get_saved_configurations(self):
        r = self.mvcm.get("/saved-configurations", "application/json")
        return r.json() if r.ok else []
    
    def download_configuration(self, config_name, download_location):
        r = self.mvcm.getzip('/saved-configurations/' + config_name, 'zip')
        if r.status_code == 200:
            with open(download_location, 'wb') as f:
                f.write(r.content)
            return True
        else:
            return False

    def restore_configuration(self, config_name):
        response = self.mvcm.post(f'/saved-configurations/{config_name}/operations/restore', None)
        if not response.ok:
            printResponseError(response)
        return response.ok

    def upload_configuration(self, file_path):
        filename = os.path.basename(file_path)
        response = self.mvcm.postbinary('/saved-configurations', file_path)
        return response.ok

    def create_configuration(self, name, description):
        data = {"name": name, "description": description if description else None}
        response = self.mvcm.post('/saved-configurations/' + name, data)
        return response.ok

    def update_configuration(self, source_hostname, target_hostname, username, password):
        try:
            # Create a temporary directory for the merge process
            with tempfile.TemporaryDirectory() as merge_base_dir:
                # Create subdirectories for extraction and merging
                source_extract_dir = os.path.join(merge_base_dir, "sourceExtracted")
                target_extract_dir = os.path.join(merge_base_dir, "targetExtracted")
                merge_file_dir = os.path.join(merge_base_dir, "mergeFile")
                
                os.makedirs(source_extract_dir, exist_ok=True)
                os.makedirs(target_extract_dir, exist_ok=True)
                os.makedirs(merge_file_dir, exist_ok=True)
                
                # Create MVCM connections
                source_connection = mvcm.Mvcm()
                target_connection = mvcm.Mvcm()
                source_connection.connect(source_hostname, username, password)
                target_connection.connect(target_hostname, username, password)
                
                # Try to delete existing configurations on servers
                try:
                    source_connection.delete('/saved-configurations/source_Merge')
                except Exception:
                    pass
                    
                try:
                    target_connection.delete('/saved-configurations/target_Merge')
                except Exception:
                    pass
                
                # Create new configurations
                source_data = {"name": 'source_Merge', "description": 'Newly created source config to be merged'}
                target_data = {"name": 'target_Merge', "description": 'Newly created target config to be merged'}
                
                source_connection.post('/saved-configurations/source_Merge', source_data)
                target_connection.post('/saved-configurations/target_Merge', target_data)

                # Download configurations
                source_response = source_connection.getzip('/saved-configurations/source_Merge', 'zip')
                target_response = target_connection.getzip('/saved-configurations/target_Merge', 'zip')
                
                # Save downloaded zip files in the temporary directory
                source_zip = os.path.join(merge_base_dir, "source_Merge.zip")
                target_zip = os.path.join(merge_base_dir, "target_Merge.zip")
                
                with open(source_zip, 'wb') as f:
                    f.write(source_response.content)
                with open(target_zip, 'wb') as f:
                    f.write(target_response.content)
                    
                # Merge configurations using the temporary directories
                merged_zip_path = source_connection.merge_configurations(
                    username, source_hostname, target_hostname,
                    merge_base_dir, source_extract_dir, target_extract_dir, merge_file_dir
                )
                
                # Upload merged configuration if the merged zip file exists
                success = False
                if os.path.exists(merged_zip_path):
                    response = target_connection.postbinary('/saved-configurations', merged_zip_path)
                    success = response.ok
                    
                # Temporary files and directories are automatically removed here
                return success

        except Exception as e:
            print(f"Error in update_configuration: {str(e)}")
            return False
        
# -------------------------------
# Excel Parser
# -------------------------------
import pandas as pd
from openpyxl import load_workbook
import numpy as np

class ExcelParser:
    def __init__(self):
        self.current_data = None
        self.current_headers = None
        self.current_file = None
    
    def read_excel(self, file_path, sheet_name=0):
        """
        Reads an Excel file and returns headers and data.
        
        Args:
            file_path (str): Path to the Excel file
            sheet_name (str or int): Sheet name or index to read (default is first sheet)
            
        Returns:
            tuple: (headers, data) where headers is a list of column names and 
                  data is a list of tuples representing rows
        """
        self.current_file = file_path
        
        try:
            # Read Excel file
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Clean up data
            df = self._clean_dataframe(df)
            
            # Store for later use
            self.current_data = df
            self.current_headers = df.columns.tolist()
            
            # Convert to format suitable for treeview
            data_rows = [tuple(row) for row in df.values]
            
            return self.current_headers, data_rows
        
        except Exception as e:
            print(f"Error reading Excel file: {str(e)}")
            raise
    
    def _clean_dataframe(self, df):
        """
        Clean up the dataframe by:
        1. Removing empty rows and columns
        2. Converting NaN values to empty strings
        3. Ensuring column names are strings and unique
        """
        # Drop completely empty rows and columns
        df = df.dropna(how='all').dropna(axis=1, how='all')
        
        # Ensure column names are strings and there are no duplicates
        df.columns = [f"Column {i}" if pd.isna(col) else str(col) for i, col in enumerate(df.columns)]
        
        # Handle duplicate column names by adding numbers
        counts = {}
        new_columns = []
        for col in df.columns:
            if col in counts:
                counts[col] += 1
                new_columns.append(f"{col}_{counts[col]}")
            else:
                counts[col] = 0
                new_columns.append(col)
        df.columns = new_columns
        
        # Replace NaN with empty strings
        df = df.fillna('')
        
        return df
    
    def get_sheet_names(self, file_path):
        """
        Returns a list of sheet names in the Excel file
        """
        try:
            workbook = load_workbook(file_path, read_only=True)
            return workbook.sheetnames
        except Exception as e:
            print(f"Error getting sheet names: {str(e)}")
            raise
    
    def get_json_data(self):
        """
        Returns the current data as a JSON-serializable list of dictionaries
        """
        if self.current_data is None:
            return None
        
        # Convert DataFrame to list of dicts, handling non-serializable objects like numpy numbers
        records = self.current_data.to_dict('records')
        
        # Ensure all values are JSON-serializable and exclude specific keys
        for record in records:          
            for key, value in record.items():
                if isinstance(value, np.integer):
                    record[key] = int(value)
                elif isinstance(value, np.floating):
                    record[key] = float(value)
                elif isinstance(value, np.ndarray):
                    record[key] = value.tolist()
        
        return records
    
    def extract_names(self, json_list):
        """
        Iterates through each JSON object in the list, extracts the 'name' value,
        and returns a list of these names.
        """
        names = []
        for json_obj in json_list:
            name = json_obj.get('name')
            if name:
                names.append(name)
        return names


    
    def save_to_excel(self, file_path, data=None, headers=None):
        """
        Save data to an Excel file
        
        Args:
            file_path (str): Path to save the Excel file
            data (list or DataFrame): Data to save. If None, uses current_data
            headers (list): Column headers. If None, uses current_headers
        """
        try:
            if data is not None:
                if isinstance(data, pd.DataFrame):
                    df = data
                else:
                    # Convert list of tuples/lists to DataFrame
                    df = pd.DataFrame(data, columns=headers or self.current_headers)
            else:
                df = self.current_data
            
            if df is not None:
                df.to_excel(file_path, index=False)
                return True
            else:
                raise ValueError("No data to save")
        
        except Exception as e:
            print(f"Error saving to Excel: {str(e)}")
            raise

    def create_ccs_server(self, mvcm_inst, json_list):
        """
        Iterates through each JSON object, extracts the 'name' value, attaches it to the URL,
        removes it from the JSON, and passes the modified JSON into the post function.
        """
        for json_obj in json_list:
            # Extract the 'name' value and remove it from the JSON object
            name = json_obj.pop('name', None)
            serverName = json_obj.pop('Chip', None)
            serverName = json_obj.pop('Port', None)
            serverName = json_obj.pop('Location', None)
            
            if name is not None:
                # Construct the URL with the 'name' value
                url = f'/ccs/servers/{name}'
                
                # Pass the modified JSON object into the post function
                response = mvcm_inst.post(url, json_obj)
                print(printResponseError(response))
                # Check if the response is OK
                if not response.ok:
                    print(f"Failed to create CCS server for {name}")
                    return False
            else:
                print("No 'name' key found in JSON object")
                return False
        
        return True
    
    def create_ccs_console(self, mvcm_inst, json_list):
        """
        Iterates through each JSON object, extracts the 'name' value, attaches it to the URL,
        removes it from the JSON, and passes the modified JSON into the post function.
        """
        for json_obj in json_list:
            # Extract the 'name' value and remove it from the JSON object
            serverName = json_obj.pop('Chip', None)
            sessionName = json_obj.pop('name', None)
            DR = json_obj.pop('DR?', None)
            LPAR = json_obj.pop('LPAR', None)
            CUAddress = json_obj.pop('CU Address', None)
            
            if serverName is not None and sessionName is not None:
                # Construct the URL with the 'name' value
                url = f'/ccs/servers/{serverName}/sessions/{sessionName}'
                
                # Pass the modified JSON object into the post function
                response = mvcm_inst.post(url, json_obj)
                # Check if the response is OK
                if not response.ok:
                    print(f"Failed to create CCS session {sessionName} for {serverName}")
                    return False
            else:
                print("No 'name' key found in JSON object")
                return False
        
        return True
